
import base64
import json
import re
import time
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

EMAIL_REGEX = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.IGNORECASE)

KC_PRIMARY = "#B04EF0"
KC_ACCENT = "#E060F0"
KC_DEEP = "#8030F0"
KC_SOFT = "#F6F0FF"


def _get_secret(path: List[str], default=None):
    cur = st.secrets
    for key in path:
        if key not in cur:
            return default
        cur = cur[key]
    return cur


def _inject_brand_css():
    st.markdown(
        f"""
        <style>
          .kc-hero {{
            padding: 18px 18px;
            border-radius: 18px;
            background: linear-gradient(90deg, {KC_DEEP} 0%, {KC_PRIMARY} 45%, {KC_ACCENT} 100%);
            color: white;
            box-shadow: 0 10px 30px rgba(0,0,0,0.08);
          }}
          .kc-hero h1 {{
            margin: 0;
            font-size: 28px;
            line-height: 1.2;
          }}
          .kc-hero p {{
            margin: 6px 0 0 0;
            opacity: 0.95;
            font-size: 14px;
          }}
          .kc-card {{
            background: white;
            border: 1px solid rgba(176, 78, 240, 0.18);
            border-radius: 16px;
            padding: 14px 14px;
            box-shadow: 0 10px 24px rgba(20, 6, 31, 0.04);
          }}
          .kc-muted {{
            color: rgba(20, 6, 31, 0.72);
          }}
          div.stButton > button {{
            border-radius: 14px !important;
            border: 0 !important;
            background: linear-gradient(90deg, {KC_DEEP} 0%, {KC_PRIMARY} 55%, {KC_ACCENT} 100%) !important;
            color: white !important;
            padding: 0.55rem 1rem !important;
            font-weight: 600 !important;
          }}
          section[data-testid="stFileUploaderDropzone"] {{
            border-radius: 14px;
            border: 2px dashed rgba(176, 78, 240, 0.35);
            background: {KC_SOFT};
          }}
          div[data-testid="stDataFrame"] {{
            border-radius: 14px;
            overflow: hidden;
          }}
          .block-container {{
            padding-top: 1.2rem;
          }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def _logo_html(width_px: int = 220, top_pad_px: int = 8) -> str:
    logo_path = Path(__file__).parent / "assets" / "KrispCallLogo.png"
    if not logo_path.exists():
        return ""
    b64 = base64.b64encode(logo_path.read_bytes()).decode("utf-8")
    return f"""
      <div style="padding-top:{top_pad_px}px;">
        <img src="data:image/png;base64,{b64}" style="width:{width_px}px; height:auto;" />
      </div>
    """


def require_login():
    st.session_state.setdefault("authenticated", False)
    if st.session_state["authenticated"]:
        return

    _inject_brand_css()

    c1, c2 = st.columns([1, 2])
    with c1:
        st.markdown(_logo_html(width_px=260, top_pad_px=14), unsafe_allow_html=True)
    with c2:
        st.markdown(
            """
            <div class="kc-hero">
              <h1>Payment Summary</h1>
              <p>Secure login required. Credentials are stored in Streamlit Secrets.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.write("")
    with st.form("login_form", clear_on_submit=False):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login")

    if not submitted:
        st.stop()

    expected_user = _get_secret(["auth", "username"])
    expected_pass = _get_secret(["auth", "password"])

    if expected_user is None or expected_pass is None:
        st.error("Missing auth secrets. Add the [auth] section to Streamlit Secrets.")
        st.stop()

    if username == expected_user and password == expected_pass:
        st.session_state["authenticated"] = True
        st.rerun()
    else:
        st.error("Invalid credentials.")
        st.stop()


def _basic_auth_header(user: str, password: str) -> str:
    token = base64.b64encode(f"{user}:{password}".encode("utf-8")).decode("utf-8")
    return f"Basic {token}"


def _mixpanel_headers() -> Dict[str, str]:
    auth_header_value = _get_secret(["mixpanel", "authorization"], default=None)
    if auth_header_value:
        return {"accept": "text/plain", "authorization": str(auth_header_value).strip()}

    user = _get_secret(["mixpanel", "service_account_username"], default=None)
    pwd = _get_secret(["mixpanel", "service_account_password"], default=None)
    api_secret = _get_secret(["mixpanel", "api_secret"], default=None)

    if user and pwd:
        return {"accept": "text/plain", "authorization": _basic_auth_header(user, pwd)}
    if api_secret:
        return {"accept": "text/plain", "authorization": _basic_auth_header(api_secret, "")}

    raise RuntimeError("Missing Mixpanel credentials in secrets.")


@st.cache_data(show_spinner=False, ttl=60 * 10)
def fetch_mixpanel_event_export(project_id: int, base_url: str, from_date: date, to_date: date, event_name: str, timeout_s: int = 120) -> pd.DataFrame:
    url = f"{base_url.rstrip('/')}/api/2.0/export"
    params = {"project_id": project_id, "from_date": from_date.isoformat(), "to_date": to_date.isoformat(), "event": json.dumps([event_name])}
    resp = requests.get(url, params=params, headers=_mixpanel_headers(), timeout=timeout_s)
    if resp.status_code != 200:
        raise RuntimeError(f"Mixpanel export failed for '{event_name}'. Status {resp.status_code}. Body: {resp.text[:500]}")

    rows: List[Dict] = []
    for line in resp.text.splitlines():
        if not line.strip():
            continue
        obj = json.loads(line)
        props = obj.get("properties", {}) or {}
        out = {"event": obj.get("event")}
        out.update(props)
        rows.append(out)
    return pd.DataFrame(rows)


def dedupe_mixpanel_export(df: pd.DataFrame) -> pd.DataFrame:
    required = ["event", "distinct_id", "time", "$insert_id"]
    if any(c not in df.columns for c in required):
        return df
    df = df.copy()

    t = pd.to_numeric(df["time"], errors="coerce")
    if t.notna().all():
        if float(t.median()) > 1e11:
            t = (t // 1000)
        df["_time_s"] = t.astype("Int64")
    else:
        dt = pd.to_datetime(df["time"], errors="coerce", utc=True)
        df["_time_s"] = (dt.view("int64") // 10**9).astype("Int64")

    df = df.sort_values(["_time_s"], kind="mergesort")
    df = df.drop_duplicates(subset=["event", "distinct_id", "_time_s", "$insert_id"], keep="last").drop(columns=["_time_s"])
    return df


def _extract_all_emails(value) -> List[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    s = str(value).strip()
    if not s:
        return []
    found = EMAIL_REGEX.findall(s)
    out, seen = [], set()
    for e in found:
        e2 = e.strip().lower()
        if e2 and e2 not in seen:
            seen.add(e2)
            out.append(e2)
    return out


def validate_email_cell(value) -> bool:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return True
    s = str(value).strip()
    if not s:
        return True
    return len(_extract_all_emails(s)) > 0


def find_email_columns(df: pd.DataFrame) -> List[str]:
    return [c for c in df.columns if "email" in c.lower()]


def coalesce_email(df: pd.DataFrame, email_cols: List[str]) -> pd.Series:
    if not email_cols:
        return pd.Series([None] * len(df), index=df.index, dtype="object")

    def first_email(val):
        emails = _extract_all_emails(val)
        return emails[0] if emails else None

    tmp = pd.DataFrame({c: df[c].map(first_email) for c in email_cols})
    return tmp.bfill(axis=1).iloc[:, 0]


def pick_first_existing_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in df.columns:
            return c
    return None


def ensure_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0.0)


def unique_new_col(df: pd.DataFrame, base: str) -> str:
    if base not in df.columns:
        return base
    i = 1
    while f"{base}_{i}" in df.columns:
        i += 1
    return f"{base}_{i}"


def _annotate_bars_with_ints(ax):
    for container in ax.containers:
        labels = []
        for v in container.datavalues:
            try:
                labels.append(str(int(round(float(v)))))
            except Exception:
                labels.append("")
        ax.bar_label(container, labels=labels, padding=2, fontsize=8)


def build_excel_with_tables_and_chart(deals_joined: pd.DataFrame, owner_breakdown: pd.DataFrame, chart_fig, owner_breakdown_rounded: pd.DataFrame, from_date: date, to_date: date) -> Tuple[str, bytes]:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Deals_with_Payments"
    for r in dataframe_to_rows(deals_joined, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet("Owner_Breakdown")
    for r in dataframe_to_rows(owner_breakdown, index=False, header=True):
        ws2.append(r)

    ws3 = wb.create_sheet("Chart")
    img_bytes = BytesIO()
    chart_fig.savefig(img_bytes, format="png", bbox_inches="tight", dpi=150)
    img_bytes.seek(0)

    import tempfile
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
        tmp.write(img_bytes.getvalue())
        tmp_path = tmp.name

    img = XLImage(tmp_path)
    img.anchor = "A1"
    ws3.add_image(img)

    # numbers table at H1
    start_col = 8
    start_row = 1
    for r_idx, row in enumerate(dataframe_to_rows(owner_breakdown_rounded, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_col):
            ws3.cell(row=r_idx, column=c_idx, value=value)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"payment_summary_{from_date.strftime('%b%d').lower()}_{to_date.strftime('%b%d').lower()}.xlsx"
    return fname, out.getvalue()


def expand_deals_rows_for_multiple_emails(deals_df: pd.DataFrame, preferred_email_cols: List[str]) -> Tuple[pd.DataFrame, List[int], List[int]]:
    invalid_rows = set()
    for c in preferred_email_cols:
        if c in deals_df.columns:
            bad = ~deals_df[c].map(validate_email_cell)
            if bad.any():
                invalid_rows.update((deals_df.index[bad] + 1).tolist())

    expanded_rows = []
    missing_rows = []

    for idx, row in deals_df.iterrows():
        chosen_emails = []
        for c in preferred_email_cols:
            if c in deals_df.columns:
                emails = _extract_all_emails(row.get(c))
                if emails:
                    chosen_emails = emails
                    break

        if not chosen_emails:
            r = row.to_dict()
            r["email"] = None
            expanded_rows.append(r)
            missing_rows.append(idx + 1)
            continue

        for e in chosen_emails:
            r = row.to_dict()
            r["email"] = e
            expanded_rows.append(r)

    expanded_df = pd.DataFrame(expanded_rows)
    return expanded_df, sorted(invalid_rows), missing_rows


def main():
    st.set_page_config(page_title="KrispCall Payment Summary", page_icon="📈", layout="wide")
    require_login()
    _inject_brand_css()

    # Padding after login
    st.markdown('<div style="height:10px;"></div>', unsafe_allow_html=True)

    left, right = st.columns([1, 3], vertical_alignment="center")
    with left:
        st.markdown(_logo_html(width_px=240, top_pad_px=10), unsafe_allow_html=True)
    with right:
        st.markdown(
            """
            <div class="kc-hero">
              <h1>KrispCall Payment Summary</h1>
              <p>Upload a deals CSV. Pull Mixpanel payments and refunds. Export clean summaries in seconds.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.write("")

    with st.sidebar:
        st.markdown("### Setup")
        st.markdown(
            """
            <div class="kc-card">
              <div><b>Step 1</b>. Pick date range</div>
              <div><b>Step 2</b>. Upload deals CSV</div>
              <div><b>Step 3</b>. Run and download</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.write("")
        from_date = st.date_input("Date from", value=date.today().replace(day=1))
        to_date = st.date_input("Date to", value=date.today())
        st.write("")
        st.markdown(
            """
            <div class="kc-card kc-muted">
              Configure login and Mixpanel authorization in Streamlit Secrets.
            </div>
            """,
            unsafe_allow_html=True,
        )

    if from_date > to_date:
        st.error("Date from must be on or before Date to.")
        st.stop()

    st.markdown('<div class="kc-card">', unsafe_allow_html=True)
    deals_file = st.file_uploader("Upload deals CSV", type=["csv"])
    run = st.button("Run", type="primary", disabled=(deals_file is None))
    st.markdown("</div>", unsafe_allow_html=True)

    if not run:
        st.stop()

    with st.spinner("Working on it..."):
        deals_raw = pd.read_csv(deals_file)
        preferred_deal_email_cols = ["Person - Email - Other", "Person - Email - Work", "Person - Email - Home", "email"]

        deals_df, invalid_rows, missing_rows = expand_deals_rows_for_multiple_emails(deals_raw, preferred_deal_email_cols)

        project_id = int(_get_secret(["mixpanel", "project_id"]))
        base_url = _get_secret(["mixpanel", "base_url"], default="https://data-eu.mixpanel.com")

        try:
            payments_df = fetch_mixpanel_event_export(project_id, base_url, from_date, to_date, "New Payment Made")
            time.sleep(0.4)
            refunds_df = fetch_mixpanel_event_export(project_id, base_url, from_date, to_date, "Refund Granted")
        except Exception as e:
            st.error("Mixpanel request failed.")
            st.code(str(e))
            st.stop()

        payments_df = dedupe_mixpanel_export(payments_df)
        refunds_df = dedupe_mixpanel_export(refunds_df)

        payments_df["email"] = coalesce_email(payments_df, find_email_columns(payments_df))
        refunds_df["email"] = coalesce_email(refunds_df, find_email_columns(refunds_df))

        pay_amount_col = pick_first_existing_column(payments_df, ["Amount", "Amount Paid", "Payment Amount", "amount"])
        ref_amount_col = pick_first_existing_column(refunds_df, ["Refund Amount", "Refunded Transaction Amount", "Refund.Amount", "refund_amount"])
        if pay_amount_col is None:
            st.error("Could not find a payment amount column in 'New Payment Made' export.")
            st.stop()
        if ref_amount_col is None:
            st.error("Could not find a refund amount column in 'Refund Granted' export.")
            st.stop()

        payments_df[pay_amount_col] = ensure_numeric(payments_df[pay_amount_col])
        refunds_df[ref_amount_col] = ensure_numeric(refunds_df[ref_amount_col])

        payments_summary = payments_df.dropna(subset=["email"]).groupby("email", as_index=False)[pay_amount_col].sum().rename(columns={pay_amount_col: "Total_Amount"})
        refunds_summary = refunds_df.dropna(subset=["email"]).groupby("email", as_index=False)[ref_amount_col].sum().rename(columns={ref_amount_col: "Refund_Amount"})

        summary = payments_summary.merge(refunds_summary, on="email", how="outer")
        summary["Total_Amount"] = ensure_numeric(summary.get("Total_Amount"))
        summary["Refund_Amount"] = ensure_numeric(summary.get("Refund_Amount"))
        summary["Net_Amount"] = summary["Total_Amount"] - summary["Refund_Amount"]

        total_col = unique_new_col(deals_df, "Total_Amount")
        refund_col = unique_new_col(deals_df, "Refund_Amount")
        net_col = unique_new_col(deals_df, "Net_Amount")

        deals_joined = deals_df.merge(summary, on="email", how="left").rename(columns={"Total_Amount": total_col, "Refund_Amount": refund_col, "Net_Amount": net_col})
        for c in [total_col, refund_col, net_col]:
            deals_joined[c] = pd.to_numeric(deals_joined[c], errors="coerce").fillna(0.0)

        owner_col = "Deal - Owner" if "Deal - Owner" in deals_joined.columns else None
        if owner_col is None:
            exact_owner = [c for c in deals_joined.columns if c.strip().lower() == "owner"]
            if exact_owner:
                owner_col = exact_owner[0]
        if owner_col is None:
            candidates = [c for c in deals_joined.columns if "owner" in c.lower()]
            owner_col = candidates[0] if candidates else None
        if owner_col is None:
            owner_col = unique_new_col(deals_joined, "Deal - Owner")
            deals_joined[owner_col] = "Unknown"

        owner_breakdown = deals_joined.groupby(owner_col, as_index=False)[[total_col, refund_col, net_col]].sum(numeric_only=True).rename(columns={owner_col: "Deal - Owner"}).sort_values(net_col, ascending=False)

        plot_df = owner_breakdown.set_index("Deal - Owner")[[total_col, refund_col, net_col]]
        fig, ax = plt.subplots()
        plot_df.plot(kind="bar", ax=ax, color=[KC_PRIMARY, KC_ACCENT, KC_DEEP])
        ax.set_xlabel("Deal - Owner")
        ax.set_ylabel("Amount")
        ax.set_title("Payment, Refund, Net by Deal - Owner")
        _annotate_bars_with_ints(ax)

        owner_breakdown_rounded = owner_breakdown.copy()
        for c in [total_col, refund_col, net_col]:
            owner_breakdown_rounded[c] = owner_breakdown_rounded[c].round(0).astype(int)

        fname, excel_bytes = build_excel_with_tables_and_chart(deals_joined, owner_breakdown, fig, owner_breakdown_rounded, from_date, to_date)

    tab1, tab2, tab3, tab4 = st.tabs(["Summary", "Deals join", "Owner breakdown", "Downloads"])
    with tab1:
        st.markdown('<div class="kc-card">', unsafe_allow_html=True)
        st.subheader("Summary by Email")
        st.dataframe(summary.sort_values("Net_Amount", ascending=False), use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with tab2:
        st.markdown('<div class="kc-card">', unsafe_allow_html=True)
        st.subheader("Deals with Payment Columns")
        st.caption(f"Deals expanded from {len(deals_raw):,} to {len(deals_df):,} rows after splitting comma-separated emails in a chosen email cell.")
        st.dataframe(deals_joined, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with tab3:
        st.markdown('<div class="kc-card">', unsafe_allow_html=True)
        st.subheader("Breakdown by Deal - Owner")
        st.dataframe(owner_breakdown, use_container_width=True)
        st.pyplot(fig, clear_figure=False)
        st.markdown("</div>", unsafe_allow_html=True)

    with tab4:
        st.markdown('<div class="kc-card">', unsafe_allow_html=True)
        st.subheader("Downloads")
        st.download_button("Download summary CSV", data=summary.to_csv(index=False).encode("utf-8"), file_name="summary_by_email.csv", mime="text/csv")
        st.download_button("Download deals join CSV", data=deals_joined.to_csv(index=False).encode("utf-8"), file_name="deals_with_payment_columns.csv", mime="text/csv")
        st.download_button("Download owner breakdown CSV", data=owner_breakdown.to_csv(index=False).encode("utf-8"), file_name="owner_breakdown.csv", mime="text/csv")
        st.download_button("Download Excel (3 tabs)", data=excel_bytes, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.markdown("</div>", unsafe_allow_html=True)

    if invalid_rows or missing_rows:
        st.write("")
        st.markdown('<div class="kc-card">', unsafe_allow_html=True)
        st.subheader("Data quality logs")
        if invalid_rows:
            with st.expander(f"Invalid email cells (non-empty but no valid email): {len(invalid_rows)} row(s)"):
                st.text("\n".join([f"Row #{n}" for n in invalid_rows[:500]]))
        if missing_rows:
            with st.expander(f"Missing email after coalescing: {len(missing_rows)} row(s)"):
                st.text("\n".join([f"Row #{n}" for n in missing_rows[:500]]))
        st.markdown("</div>", unsafe_allow_html=True)


if __name__ == "__main__":
    main()
